import requests
import craigslist_data as cr
import sys
import os.path
from pathlib import Path
# Import dattime to get now time and put in excel file
import datetime
import time
import string
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook 
from openpyxl import Workbook
from bs4 import BeautifulSoup as bs4
import datetime
from random import randint
from time import sleep
import logging

def getURLList(count,url_base):
	#url_base = 'http://sfbay.craigslist.org/search/nby/apa'
	url_list=[]
	#url_list.append(url_base)
	x=120
	while(int(count) > x):
		url_add="?s="+str(x)
		x=x+120
		url_list.append(url_base+url_add)
		print (url_base+url_add)
		logging.info(url_base+url_add)
	return url_list
def main1():
	getURLList(500)

def setLog(logfile):
	#for handler in logging.root.handlers[:]:
    #logging.root.removeHandler(handler)
	#logfile=datetime.now().strftime('mylogfile_%H_%M_%d_%m_%Y.log')
	logging.basicConfig(filename=logfile,level=logging.DEBUG,format="%(asctime)s:%(levelname)s:%(message)s")
	logging.debug('This message should go to the log file '+logfile)
	print('This message should go to the log file '+logfile)
	logging.info('So should this')
	logging.warning('And this, too')

def getFilename(argv):
	url_base =  str(sys.argv[1])
	eby = 'eby'
	nby = 'nby'
	sby = 'sby'
	prefixName=""
	if eby in url_base:
	    prefixName=eby
	if nby in url_base:
	    prefixName=nby
	if sby in url_base:
	    prefixName=sby	
	now = datetime.datetime.now()
	my_filename=prefixName+"_"+now.strftime("%Y_%m_%d_%H_%M")
	return my_filename	


def main(argv):
    # my code here
	#url_base = 'http://sfbay.craigslist.org/search/nby/apa'
	url_base =  str(sys.argv[1])
	xl_file =  str(sys.argv[2])
	print("Argument passed:",url_base)
	logging.info("Argument passed:",url_base)
	logging.info("Argument passed for output:",xl_file)
	rsp=cr.getRespByURL(url_base)
	
	

	now = datetime.datetime.now()
	#print(rsp.url)
	count=cr.getTotalCount(rsp)
	print(count)
	logging.info("Total records:",count)
	
	base_filename=getFilename(sys.argv)
	my_filename=base_filename+".xlsx"
	setLog(base_filename+".log")
	mypath = Path().absolute()
	print("Current path:",mypath)
	logging.info("Current path:",mypath)
	#my_filenameWithPath =  "/Users/sumeet_badwal/"+my_filename
	#my_filenameWithPath =  "C:\\Users\\sumeet_badwal\\Documents\\"+my_filename
	#my_filenameWithPath =  "C:\\Users\\sumeet_badwal\\Downloads\\"+my_filename
	my_filenameWithPath =  my_filename
	if(os.path.isfile(my_filenameWithPath) ):
		wb=load_workbook(my_filename)
	else:
		wb = Workbook()

	# grab the active worksheet
	ws = wb.active
	resultList=getURLList(count,url_base)
	for resultURL in resultList:
		#rsp=cr.getRespByURL(resultURL)
		data_list=cr.getResultsAndSave(resultURL)
		# will sleep a random number of seconds (between 10 and 100).
		sleep(randint(10,100))
		#cr.addToWB(ws,wb,my_filename,link_all, name_all, price_all, bed_bath_desc_all, sq_ft_all, location_all, utilities_all, avaliable_date_all, desc_all)
		cr.addToWB(ws,wb,my_filename,data_list)#link_all, name_all, price_all, bed_bath_desc_all, sq_ft_all, location_all, utilities_all, avaliable_date_all, desc_all)
if __name__ == "__main__":
    main(sys.argv[1:])


