import requests

#url_based = 'http://sfbay.craigslist.org/search/nby/apa'
#params = dict(bedrooms=1, is_furnished=1)
# Define our URL and a query we want to post
#url='https://sfbay.craigslist.org/search/nby/apa?search_distance=1&postal=95124&min_bedrooms=3&minSqft=1400&availabilityMode=0'
#base_url='https://sfbay.craigslist.org'

#url = base_url + 'search/eby/apa?nh=48&anh=49&nh=112&nh=58&nh=61&nh=62&nh=66&min_price=2200&bedrooms=1'


#rsp = requests.get(url, params=params)
def getRespByURL(url_base):
	rsp = requests.get(url_base)#(url)
	#print(rsp.url)
	# We can access the content of the response that Craigslist sent back here:
	print(rsp.text[:500].encode('utf-8'))
	return rsp


import sys
import os.path
# Import dattime to get now time and put in excel file
import datetime
import time
import string
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook 
from openpyxl import Workbook
from bs4 import BeautifulSoup as bs4
import house

		
def getTotalCount(rsp):
	html = bs4(rsp.text, 'html.parser')
	total_count = html.find_all('span', attrs={'class': 'totalcount'})
	if(len(total_count) > 0):
		return total_count[0].text
	else:
		return 0

# BS4 can quickly parse our text, make sure to tell it that you're giving html
def getHTMLByResp(rsp):
	html = bs4(rsp.text, 'html.parser')
	
	# BS makes it easy to look through a document
	#print(html.prettify()) #[:3000])
	# find_all will pull entries that fit your search criteria.
	# Note that we have to use brackets to define the `attrs` dictionary
	# Because "class" is a special word in python, so we need to give a string.
	apts = html.find_all('p', attrs={'class': 'result-info'})
	total_count = html.find_all('span', attrs={'class': 'totalcount'})
	if(len(total_count) > 0):
		print (total_count[0].text)
	getResultsAndSave()
	print (len(apts))
	
def getAptCount(apts):
	print (len(apts))
	return len(apts)


def getPriceFromTextTried(price_apt,txt):
	posting_list=[]
	if(len(price_apt) == 0 ):
		price_alt_apt_list = txt.find_all('span', attrs={'class': 'postingtitletext'})
		
		if(len(price_alt_apt_list) > 0 ):
			price_apt_alt = price_alt_apt_list[0].find_all('span', attrs={'class': 'price'}) 
			housing_apt_list = price_alt_apt_list[0].find_all('span', attrs={'class': 'housing'})
			if(len(price_apt_alt) > 0):
				price_alt_soup = BeautifulSoup(str(price_apt_alt[0]), 'html.parser')
				# price_all.append(price_alt_soup.text)
				return price_alt_soup.text
			else:
				return "1 Price not found"
		else:
			return "2 Price not found"
	else:
		return "3 Price not found"   
		
def getPriceFromText(price_apt,txt):
	# Extract it to different method
	if(len(price_apt) == 0 ):
		price_alt_apt_list = txt.find_all('span', attrs={'class': 'postingtitletext'})
		if(len(price_alt_apt_list) > 0 ):
			price_apt_alt = price_alt_apt_list[0].find_all('span', attrs={'class': 'price'}) 
			if(len(price_apt_alt) > 0):
				price_alt_soup = BeautifulSoup(str(price_apt_alt[0]), 'html.parser')
				# price_all.append(price_alt_soup.text)
				return price_alt_soup.text
			else:
				return "1 Price not found"
		else:
			return "2 Price not found"
	else:
		return price_apt[0].text   
		
def getBedBathFromText(price_apt,txt):
	price_alt_apt_list = txt.find_all('span', attrs={'class': 'postingtitletext'})
	if(len(price_alt_apt_list) > 0 ):
		price_apt_alt = price_alt_apt_list[0].find_all('span', attrs={'class': 'housing'}) 
		if(len(price_apt_alt) > 0):
			price_alt_soup = BeautifulSoup(str(price_apt_alt[0]), 'html.parser')
			# price_all.append(price_alt_soup.text)
			return price_alt_soup.text
		else:
			return "1 Bed Bath not found"
	else:
		return "2 Bed Bath not found"
	

def getResultsAndSave(url_base):
	
	data_list=[]
	house_list=[]

	# This will remove weird characters that people put in titles like ****!***!!!
	use_chars = string.ascii_letters + ''.join([str(i) for i in range(10)]) + ' '


	link_list = []  # We'll store the data here

	resp = requests.get(url_base)#(url)
	txt = bs4(resp.text, 'html.parser')
	apts = txt.findAll(attrs={'class': "result-info"})

	# We're just going to pull the title and link
	for apt in apts:#[:2]:
		house_ins = None
		try:
			price_apt = apt.find_all('span', attrs={'class': 'result-price'})
			price_soup = BeautifulSoup(str(price_apt), 'html.parser')
			print (price_soup.text)
			title = apt.find_all('a', attrs={'class': 'result-title hdrlnk'})[0]
			#title_soup = BeautifulSoup(title.text, 'html.parser')
			name = ''.join([i for i in title.text if i in use_chars])
			link = title.attrs['href']
			#link_all.append([link])
			#name_all.append([name])
			# Check if there is not result price yet, get it from different way
			#if(len(price_apt) > 0 ):
			#	price_all.append([price_soup.text])
			
			print (name) 
			print (price_apt)
			if link not in link_list:
				url_each = ""
				#url_each = base_url+link 
				url_each=url_base+link
				if(link.startswith("http")):
					url_each = link
				print (url_each)
				house_ins = house.house(link)
				house_ins.setTitle(name)
				
				# Open each link, get its response  and parse it
				resp_each = requests.get(url_each)
				txt = bs4(resp_each.text, 'html.parser')
				# Get it from different method
				price_found=getPriceFromText(price_apt,txt)
				house_ins.setPrice(price_found)
				#price_all.append(price_found)   
				txt_attr = txt.find_all('p', attrs={'class': 'attrgroup'})
				txt_attr_utils=[]
				bedbathsoup=""
				sqftsoup=""
				availabledatesoup=""
				location=[]
				descr=""	
				for att_utils in txt_attr:
					util_each=att_utils.find_all('span')
					utilitiessoup = BeautifulSoup(str(util_each), 'html.parser')
					#print "55555"
					txt_attr_utils.append(utilitiessoup.text)
				#print (1111)
				house_ins.setUtilities(''.join(txt_attr_utils))
				for utils in txt_attr_utils:
					utilList = utils.split(",")
					
					for util in utilList:
						print(util)
						if("BR / " in util):
							bedbathsoup=util;
							
						elif("ft2" in util):
							sqftsoup=util
						elif("available" in util):
							availabledatesoup=util
				if(len(availabledatesoup) == 0):
					availabledatesoup="No available date"
				house_ins.setAvailability(availabledatesoup) # availabledatesoup
				if(len(sqftsoup) == 0):
					sqftsoup="No sqft found"
				house_ins.setSqft(sqftsoup)
				if(len(bedbathsoup) == 0):
					bedbathsoup="No bed bath found"
				house_ins.setBedBath(bedbathsoup)
				locations=txt.find_all('div', attrs={'id': 'map'})
				#print(locations)
				#house_ins.setBedBath("BedBath")
				if(len(locations) > 0):
					house_ins.setLatitutde(locations[0].get('data-latitude'))
					house_ins.setLongitude(locations[0].get('data-longitude'))
				else:
					house_ins.setLatitutde("No latitude")
					house_ins.setLongitude("No longitude")
				desc_all = txt.find_all('meta',attrs={'property': 'og:description'})#,limit=1)
				desc=[]
				for desc_each in desc_all:
					desc.append(desc_each.get('content'))
				print("******")
				if(len(desc) > 0):
					house_ins.setDesc(desc)
				else:
					house_ins.setDesc("No Desc found")
				
				descr=desc[0]
				
		except Exception as e:
			print("Error occured",e)
			pass
		if house_ins is None:
			pass
		else:
			house_list.append(house_ins)

	return house_list
	
def getAvailableDate(txt):
	print ("@22")
	
    
def printLinksCount(link_all):
	print(len(link_all))

def getDataListInd(ind):
	ind=ind+1
	return data_list[ind]	
		
def addToWB(ws,wb,my_filename,data_list):#,link_all, name_all, price_all, bed_bath_desc_all, sq_ft_all, location_all, utilities_all, avaliable_date_all, desc_all):
	index=0
	existrowcount=1
	ind=0;
	
	# Get existing row count from the file. This is needed to start putting new links after last row count.
	for row in ws.iter_rows():
		existrowcount = existrowcount+1
	print(existrowcount)

	# Loop over found link and append to excel file
	for house in data_list:
		#ws.append([str(link_all[index]),str(name_all[index]), str(price_all[index]),str(bed_bath_desc_all[index]),str(sq_ft_all[index]),str(location_all[index]),str(utilities_all[index]) ,str(avaliable_date_all[index]),str(desc_all[index]),datetime.datetime.now()])
		ws.append([str(house.getLink()),str(house.getTitle()), str(house.getPrice()),str(house.getBedBath()),str(house.getSqft()),str(house.getLatitude()),str(house.getLongitude()), str(house.getUtils()) ,str(house.getAvailablity()),str(house.getDesc()),datetime.datetime.now()])
		#index=index+1
		


	# Save the existing file
	wb.save(my_filename)

def addToWBOrig(ws,wb,my_filename,link_all, name_all, price_all, bed_bath_desc_all, sq_ft_all, location_all, utilities_all, avaliable_date_all, desc_all):
	index=0
	existrowcount=1
	# Get existing row count from the file. This is needed to start putting new links after last row count.
	for row in ws.iter_rows():
		existrowcount = existrowcount+1
	print(existrowcount)

	# Loop over found link and append to excel file
	for rowNum in link_all:
		ws.append([str(link_all[index]),str(name_all[index]), str(price_all[index]),str(bed_bath_desc_all[index]),str(sq_ft_all[index]),str(location_all[index]),str(utilities_all[index]) ,str(avaliable_date_all[index]),str(desc_all[index]),datetime.datetime.now()])
		index=index+1
		


	# Save the existing file
	wb.save(my_filename)



